"""Генерация xlsx-реестра с нуля в формате СИБУР."""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from matcher import ReestrRow

log = logging.getLogger(__name__)


def _format_date(date_str: str) -> str:
    """Конвертирует '2026-04-08' → '08.04.2026'. Если уже ДД.ММ.ГГГГ — оставляет."""
    if not date_str:
        return ""
    if "." in date_str and len(date_str) == 10:
        return date_str  # уже ДД.ММ.ГГГГ
    if "-" in date_str:
        parts = date_str.split("-")
        if len(parts) == 3:
            return f"{parts[2]}.{parts[1]}.{parts[0]}"
    return date_str

# Заголовки колонок (точно как в шаблоне 44444.xlsx)
HEADERS = [
    "Контрагент",
    "Номер Акта выполненных работ",
    "Дата Акта выполненных работ",
    "Номер Счета-фактуры(Инвойса)",
    "Дата Счета-фактуры (Инвойса)",
    "Номер Договора c СИБУР",
    "Номер допсоглашения c СИБУР",
    "Наименование груза",
    "Название Пункта Отправления",
    "Наименование Грузоотправителя",
    "Название Пункта Получения",
    "Регион доставки",
    "Наименование Грузополучателя",
    "Номер накладной",
    "Дата накладной (отгрузки)",
    "Номер транспортного средства",
    "Номер полуприцепа",
    "Количество, тн",
    "Валюта",
    "Единая Ставка (без НДС)",
    "Ставка НДС",
    "Сумма НДС",
    "Поставщик услуг",
    "ИНН Поставщика",
    "Номер Счета-фактуры",
    "Дата Счета-фактуры",
    "Дата доставки (как в накладной)",
    "Номер Транспортировки",
]

# Ширина колонок (из шаблона)
COL_WIDTHS = [
    21, 12.5, 17, 17, 17, 14, 14, 35, 18, 25,
    18, 14, 25, 14, 17, 16, 14, 12, 10, 16,
    12, 12, 18, 16, 16, 17, 17, 16,
]

# Стили (точно как в шаблоне 44444.xlsx)
HEADER_FONT = Font(name="Times New Roman", size=12)
HEADER_FILL = PatternFill("solid", fgColor="E7FFB7")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Times New Roman", size=12)
DATA_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _row_to_values(row: ReestrRow) -> list:
    """Конвертирует ReestrRow в список значений для 28 колонок."""
    return [
        'ООО "РИАРРА"',                           # 1  Контрагент
        row.act_number or None,                    # 2  Номер Акта
        _format_date(row.act_date) or None,        # 3  Дата Акта
        row.invoice_number or None,                # 4  Номер СФ (Инвойса)
        _format_date(row.act_date) or None,        # 5  Дата СФ (та же дата)
        row.contract or "СХ.38122",                # 6  Договор
        None,                                      # 7  Допсоглашение
        row.product_name or None,                  # 8  Груз
        row.loading_city or None,                  # 9  Пункт Отправления
        row.sender_name or None,                   # 10 Грузоотправитель
        row.unloading_city or None,                # 11 Пункт Получения
        row.country or None,                       # 12 Регион
        row.receiver_name or None,                 # 13 Грузополучатель
        row.cmr_number or None,                    # 14 Номер накладной (CMR)
        row.loading_date or None,                  # 15 Дата накладной
        row.vehicle_number or None,                # 16 Номер ТС
        row.trailer_number or None,                # 17 Номер прицепа
        row.weight if row.weight > 0 else None,    # 18 Количество тн
        "USD",                                     # 19 Валюта
        row.rate_usd if row.rate_usd > 0 else None,  # 20 Ставка
        0,                                         # 21 Ставка НДС (всегда 0)
        0,                                         # 22 Сумма НДС (всегда 0)
        None,                                      # 23 Поставщик услуг
        None,                                      # 24 ИНН Поставщика
        None,                                      # 25 Номер СФ поставщика
        None,                                      # 26 Дата СФ поставщика
        _format_date(row.act_date) or None,        # 27 Дата доставки (= Дата Акта)
        row.transport_number or None,              # 28 Номер Транспортировки
    ]


def generate_reestr(rows: list[ReestrRow], output_path: str | Path) -> Path:
    """Генерирует чистый xlsx-реестр с нуля."""
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр"

    # Заголовки (строка 1)
    ws.row_dimensions[1].height = 166.5
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Ширина колонок
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Данные (строки 2+)
    for row_idx, row in enumerate(rows, 2):
        values = _row_to_values(row)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

    wb.save(output_path)
    log.info(f"Реестр: {output_path} ({len(rows)} строк)")
    return output_path


def generate_summary(rows: list[ReestrRow]) -> str:
    """Текстовая сводка для чата."""
    total = len(rows)
    full = sum(1 for r in rows if r.has_upd and r.has_shipment and r.has_cmr)
    no_sd = [r for r in rows if not r.has_shipment]
    no_cmr = [r for r in rows if not r.has_cmr]

    lines = [f"📋 Реестр: {total} перевозок, {full} полных"]

    if no_sd:
        lines.append(f"⚠️ Нет заявки: {', '.join(r.vehicle_number for r in no_sd)}")
    if no_cmr:
        lines.append(f"⚠️ Нет CMR: {', '.join(r.vehicle_number for r in no_cmr)}")

    total_usd = sum(r.rate_usd for r in rows)
    lines.append(f"💰 Итого: {total_usd:,.0f} USD")

    return "\n".join(lines)
